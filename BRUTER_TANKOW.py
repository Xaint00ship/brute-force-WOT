#moduls
from openpyxl import Workbook
from bs4 import BeautifulSoup as bs
from selenium.webdriver.firefox.options import Options
from selenium.webdriver import Firefox
from selenium import webdriver
import time
from datetime import date
import xlrd, xlwt
from openpyxl import load_workbook
import re 
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
Real_date = date.today()

#######
opts = Options()
opts.set_headless()
assert opts.headless

#######/getting match links
print(Real_date)
def get_hrefs_match():
	url = "https://nb-bet.com/Results"
	driver = Firefox(options=opts)############################# ПОСТАВИТЬ НАСТРОЙКИ 
	driver.get(url)
	time.sleep(4)################################# УМЕНЬШИТЬ 
	
	URLs = []
	page_number = driver.find_element_by_id('MainContent_pnlPages').find_elements_by_tag_name('a')
	if len(page_number) != 0:
		for num in range(2,len(page_number)+1,1):
			
			btn_sorted_actived = driver.find_element_by_id('MainContent_rblSort').find_elements_by_tag_name('label')[1].click()
			time.sleep(1)
			href_array = driver.find_element_by_id('MainContent_tblContent').find_elements_by_class_name('a-dotted-hover')
			for i in range(2, len(href_array) , 3):
				URLs.append(href_array[i].get_attribute("href"))
				
			btn_next_page = driver.find_element_by_id(f'MainContent_page_{num}')
			btn_next_page.click()
			time.sleep(0.3)

			if num == len(page_number):
				btn_sorted_actived = driver.find_element_by_id('MainContent_rblSort').find_elements_by_tag_name('label')[1].click()
				time.sleep(0.3)
				href_array = driver.find_element_by_id('MainContent_tblContent').find_elements_by_class_name('a-dotted-hover')
				for i in range(2, len(href_array) , 3):
					URLs.append(href_array[i].get_attribute("href"))
					
				driver.quit()
	else:
		btn_sorted_actived = driver.find_element_by_id('MainContent_rblSort').find_elements_by_tag_name('label')[1].click()
		time.sleep(1)
		href_array = driver.find_element_by_id('MainContent_tblContent').find_elements_by_class_name('a-dotted-hover')
		for i in range(2, len(href_array) , 3):
			URLs.append(href_array[i].get_attribute("href"))
		driver.quit()

	return URLs
##############################
Urls_clean_count = []
URLs = get_hrefs_match()
print(len(URLs))

for i in URLs:
	if re.search('https://nb-bet.com/LiveEvents/', i) != None:
		Urls_clean_count.append('da')
		
URLs = URLs[len(Urls_clean_count):]

print(len(URLs))
##############################

########### Seting table excel
wb = Workbook()
sheet = wb.active
sheet = wb.create_sheet('page', 0)


style_green = PatternFill(fill_type='solid',start_color='90EE90',end_color='90EE90')
style_blue = PatternFill(fill_type='solid',start_color='00BFFF',end_color='00BFFF')
style_gray_dark = PatternFill(fill_type='solid',start_color='696969',end_color='696969')
style_gray_light = PatternFill(fill_type='solid',start_color='A9A9A9',end_color='A9A9A9')

font_red = Font(name='Calibri',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF0000')
font_blue = Font(name='Calibri',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='0000FF')

align_center=Alignment(horizontal='center',vertical='bottom',text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)


##############################
sheet.cell(row=1, column=1).value ='№'
sheet.cell(row=1, column=1).fill = style_gray_dark
sheet['A1'].alignment = align_center

sheet.column_dimensions['A'].width = 7
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 20
sheet.column_dimensions['F'].width = 10
sheet.column_dimensions['G'].width = 10
sheet.column_dimensions['H'].width = 20
sheet.column_dimensions['I'].width = 20
sheet.column_dimensions['J'].width = 20

sheet.cell(row=1, column=2).value ='Чемпионат'
sheet.cell(row=1, column=2).fill = style_gray_dark
sheet['B1'].alignment = align_center

sheet.cell(row=1, column=3).value ='Дата, Время (ЕКБ)'
sheet.cell(row=1, column=3).fill = style_gray_dark
sheet['C1'].alignment = align_center

sheet.cell(row=1, column=4).value ='Команда Хозяев'
sheet.cell(row=1, column=4).fill = style_gray_dark
sheet['D1'].alignment = align_center

sheet.cell(row=1, column=5).value ='Команда Гостей '
sheet.cell(row=1, column=5).fill = style_gray_dark
sheet['E1'].alignment = align_center

sheet.cell(row=1, column=6).value ='Матчей'
sheet.cell(row=1, column=6).fill = style_gray_dark
sheet['F1'].alignment = align_center

sheet.cell(row=1, column=7).value ='Прогноз'
sheet.cell(row=1, column=7).fill = style_gray_dark
sheet['G1'].alignment = align_center

sheet.cell(row=1, column=8).value ='Вероятность исхода'
sheet.cell(row=1, column=8).fill = style_gray_dark
sheet['H1'].alignment = align_center

sheet.cell(row=1, column=9).value ='Название исхода'
sheet.cell(row=1, column=9).fill = style_gray_dark
sheet['I1'].alignment = align_center

sheet.cell(row=1, column=10).value ='Результат матча'
sheet.cell(row=1, column=10).fill = style_gray_dark
sheet['J1'].alignment = align_center

################# follow links
number_column = 0
number_math = 0

for i in range(len(URLs)):################################## ПОСТАВИТЬ len(URLs)
	if len(URLs) == 0:
		print('Ошибка, матчей нет')
	
	print(URLs[i])
	driver = Firefox(options=opts)
	driver.get(URLs[i])
	time.sleep(0.5)
	################# get count mathes

	try:
		number_of_matches_home = driver.find_elements_by_class_name('MatchTableItem__Item-sc-14xsz32-0')[1].find_elements_by_tag_name('span')[0].text
		number_of_matches_away = driver.find_elements_by_class_name('MatchTableItem__Item-sc-14xsz32-0')[1].find_elements_by_tag_name('span')[1].text
		DT = driver.find_element_by_class_name('MatchScoreboard__ScoreboardDate-sc-1al6574-3').text
		print(DT)
		btn_expense_rating = driver.find_elements_by_class_name('BorderContainer__BorderedContainer-w3yp29-0')[3].find_elements_by_tag_name('a')[0]
		btn_expense_rating.click()

		################# get match info
		time.sleep(1)
		championship_name = driver.find_element_by_id('MainContent_tblMatch').find_element_by_tag_name('a').text
		Date_Time = driver.find_element_by_id('MainContent_tblMatch').find_elements_by_tag_name('span')[1].text
		home_team = driver.find_element_by_class_name('pnl-left-team').find_elements_by_tag_name('span')[0].text
		away_team = driver.find_element_by_class_name('pnl-right-team').find_elements_by_tag_name('span')[0].text
		score = driver.find_element_by_class_name('lbl-score').text


		################## insert in exel
		number_math+=1
		print(f'матч {number_math} из {len(URLs)}')
		sheet.cell(row=number_column+2, column=1).value = number_math
		sheet.cell(row=number_column+2, column=1).fill = style_gray_light
		sheet.cell(row=number_column+2, column=1).font = font_blue
		sheet[f'A{number_column+2}'].alignment = align_center

		sheet.cell(row=number_column+2, column=2).value = championship_name
		sheet.cell(row=number_column+2, column=2).fill = style_gray_light
		sheet.cell(row=number_column+2, column=2).font = font_blue
		sheet[f'B{number_column+2}'].alignment = align_center

		sheet.cell(row=number_column+2, column=3).value = DT 
		sheet.cell(row=number_column+2, column=3).fill = style_gray_light
		sheet[f'C{number_column+2}'].alignment = align_center

		sheet.cell(row=number_column+2, column=4).value = home_team
		sheet.cell(row=number_column+2, column=4).fill = style_gray_light
		sheet.cell(row=number_column+2, column=4).font = font_blue
		sheet[f'D{number_column+2}'].alignment = align_center

		sheet.cell(row=number_column+2, column=6).value = f'{number_of_matches_home}/{number_of_matches_away}'
		sheet.cell(row=number_column+2, column=6).fill = style_gray_light
		sheet.cell(row=number_column+2, column=6).font = font_red
		sheet[f'E{number_column+2}'].alignment = align_center

		sheet.cell(row=number_column+2, column=5).value = away_team 
		sheet.cell(row=number_column+2, column=5).fill = style_gray_light
		sheet.cell(row=number_column+2, column=5).font = font_blue
		sheet[f'F{number_column+2}'].alignment = align_center

		sheet.cell(row=number_column+2, column=7).value = score
		sheet.cell(row=number_column+2, column=7).fill = style_gray_light
		sheet.cell(row=number_column+2, column=7).font = font_red
		sheet[f'G{number_column+2}'].alignment = align_center
		number_column+=1
		##################
		soup = bs(driver.page_source, 'html.parser')
		driver.quit()

		################## insert in excel
		recommendations = soup.find('div', id="MainContent_pnlDetailРекомендации").find_all('tr')
		tds_dict = {}
		for i in range(len(recommendations)):
			tds = recommendations[i].find_all('td')
			try:
				tds_dict[tds[0].text]=int(tds[1].text)
				tds_dict[tds[2].text]=int(tds[3].text)
			except:
				tds_dict[tds[0].text]=int(tds[1].text)

		list_tds = list(tds_dict.items())
		list_tds.sort(key=lambda i: i[1])
		list_tds.reverse()

		for i in list_tds:

			if i[1] >= 90:
				sheet.cell(row=number_column+1, column=8).value = i[1]
				sheet.cell(row=number_column+1, column=8).fill = style_green
				sheet[f'H{number_column+1}'].alignment = align_center

				sheet.cell(row=number_column+1, column=9).value = i[0]
				sheet.cell(row=number_column+1, column=9).fill = style_green
				sheet[f'I{number_column+1}'].alignment = align_center
			else:
				sheet.cell(row=number_column+1, column=8).value = i[1]
				sheet.cell(row=number_column+1, column=8).fill = style_blue
				sheet[f'H{number_column+1}'].alignment = align_center

				sheet.cell(row=number_column+1, column=9).value = i[0]
				sheet.cell(row=number_column+1, column=9).fill = style_blue
				sheet[f'I{number_column+1}'].alignment = align_center
			number_column+=1
		wb.save(f'../{Real_date}.xlsx')			
		tds_dict.clear()
		#################################
	except:
		driver.quit()

# get results for 1xstavka
wbs = openpyxl.load_workbook(filename = f'../{Real_date- timedelta(days=1)}.xlsx')
sheet = wbs['page']

driver = Firefox(options=opts)############################# ПОСТАВИТЬ НАСТРОЙКИ 
driver.get('https://1xstavka.ru/results/')
time.sleep(6)
window_input_game = driver.find_element_by_id('searchGames')
index = 2
count_None = 0
while True:
	if sheet[f'D{index}'].value != None:
		try:
			window_input_game.clear()
			search_name = sheet[f'D{index}'].value
			tree_sumb = f'{search_name[0]}{search_name[1]}{search_name[2]}{search_name[3]}'
			window_input_game.send_keys(tree_sumb)
			time.sleep(5)
			res = driver.find_elements_by_class_name('u-mla')[0].text
			sheet[f'J{index}'].value = res
			wbs.save(f'../{Real_date- timedelta(days=1)}.xlsx')	
		except:
			print()

	if sheet[f'H{index}'].value == None:
		count_None+=1
	index+=1
	if count_None >= len(URLs):
		break

driver.quit()



print('Программа успешно завершила работу.')
